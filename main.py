# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy import signal
from openpyxl.drawing.image import Image
from adjustText import adjust_text
from openpyxl.styles import Font
from aspose.cells import Workbook
import os
import glob


def AMPD(data):
    """
    :param data: 1-D numpy.ndarray
    :return: 波峰所在索引值的列表
    """
    p_data = np.zeros_like(data, dtype=np.int32)
    count = data.shape[0]
    arr_rowsum = []
    for k in range(1, count // 2 + 1):
        row_sum = 0
        for i in range(k, count - k):
            if data[i] > data[i - k] and data[i] > data[i + k]:
                row_sum -= 1
        arr_rowsum.append(row_sum)
    min_index = np.argmin(arr_rowsum)
    max_window_length = min_index
    #
    # if max_window_length != 0:
    #     for k1 in range(1, max_window_length + 1):
    #         for j in range(k1, count - k1):
    #             if data[j] > data[j - k1] and data[j] > data[j + k1]:
    #                 p_data[j] += 1
    #     return np.where(p_data == max_window_length)[0]
    # else:
    #     for j in range(1, count - 1):
    #         if data[j] > data[j - 1] and data[j] > data[j + 1]:
    #             p_data[j] += 1
    #     return np.where(p_data != max_window_length)[0]

    for k in range(1, max_window_length + 1):
        for i in range(k, count - k):
            if data[i] > data[i - k] and data[i] > data[i + k]:
                p_data[i] += 1
    return np.where(p_data == max_window_length)[0]


ask_continue = 0

output_sheet_name = 'walk'
output_file_name = 'walk'

# affect_side = input("請輸入患側邊編號 (1.Right/2.Left):")
# ID = input("請輸入檔名:")
# xlsx_num = int(input("請輸入分析xlsx個數:"))
# # output_file_name = input("請輸入檔名:") + '.xlsx'
#
# if not os.path.exists(ID):
#     path = os.getcwd() + '/' + ID
#     os.mkdir(ID)
#     os.mkdir(path + '/excel')
#     os.mkdir(path + '/picture')
while ask_continue < 1:
    walk = 1
    files = os.listdir(os.getcwd())
    glob_path = os.getcwd() + '/*.xlsx'
    affect_side = input("請輸入患側邊編號 (1.Right/2.Left):")
    ID = input("請輸入檔名:")

    # while True:
    xlsx_num = int(input("請輸入分析xlsx個數:"))
        # if not len(glob.glob(glob_path)) == xlsx_num:
        #     print("請再確認xlsx個數是否正確，並重新輸入")
        #     print("目前有"+str(xlsx_num) + "個xlsx")
        #     continue
        # else:
        #     break

    # output_file_name = input("請輸入檔名:") + '.xlsx'

    if not os.path.exists(ID):
        path = os.getcwd() + '/' + ID
        os.mkdir(ID)
        os.mkdir(path + '/excel')
        os.mkdir(path + '/picture')

    path = os.getcwd() + '/' + ID + '/picture/'
    path_excel = os.getcwd() + '/' + ID + '/excel/'
    # affect_pic_name = path + 'Affect_side(' + str(walk) + ').png'
    # normal_pic_name = path + 'Normal_side(' + str(walk) + ').png'
    # combine_pic_name = path + 'Combine(' + str(walk) + ').png'

    # excel_file = './knee_angle_data.xlsx'
    for xlsx_i in range(1, xlsx_num+1):
        excel_file = './knee_angle_data' + str(xlsx_i) + '.xlsx'
        affect_pic_name = path + 'Affect_side(' + str(xlsx_i) + ').png'
        normal_pic_name = path + 'Normal_side(' + str(xlsx_i) + ').png'
        combine_pic_name = path + 'Combine(' + str(xlsx_i) + ').png'

        wb = openpyxl.load_workbook(filename=excel_file)
        sheet = wb['Sheet1']
        data = pd.read_excel(excel_file, sheet_name='Sheet1')
        # print(data)

        rows_max = len(data.index)
        # print(rows_max)
        xxx = np.arange(0, rows_max)
        # print(xxx)
        yyy_L = data['knee_angle_l'].values
        yyy_R = data['knee_angle_r'].values
        # print(yyy_L[0:10])

        # affect_side = input("請輸入患側邊編號 (1.Right/2.Left):")
        step_list = []
        i = 1
        while i >= 1:
            input_text = f"請輸入第{xlsx_i}個xlsx檔第{i}組步態範圍(EX:10 20),如不再輸入則輸入n:"
            # step_input = int(input(input_text))
            step_input = input(input_text).split()
            # print(step_input)
            # print(type(step_input[0]))
            if step_input[0] == 'n':
                break
            elif len(step_input) != 2:
                print(f"輸入錯誤,請重新輸入第{i}組步態範圍")
                continue
            elif int(step_input[1]) <= int(step_input[0]):
                print('數字輸入錯誤，請重新輸入，第2個數字要大於第1個數字')
                continue
            else:
                step_input = list(map(int, step_input))
                # print(type(step_input[0]))
                step_list.append(step_input)
                i += 1

    # print(step_list)

        if affect_side == '1':

            affect_side_peak_y = []
            affect_side_peak_x = []
            for i in range(len(step_list)):
                step_range = yyy_R[step_list[i][0]:step_list[i][1]]

                affect_side_peak_y.append(round(max(step_range), 2))  # 找出peak y值並存在lsit中 採計小數點後2位
                # affect_side_peak_x.append(np.where(step_range==max(step_range)))
                affect_side_peak_x.append(step_range.tolist().index(max(step_range)) + step_list[i][0])  # 找出peak x值並存在lsit中

                # pprint(step_range)
                # pprint(affect_side_peak_y)
                # pprint(affect_side_peak_x)
            # texta=[]
            plt.plot(xxx, yyy_R, label='Affected')
            plt.xlabel('time')
            plt.ylabel('knee angle')
            plt.legend()
            plt.title('Affected side (R)')
            for i in range(len(affect_side_peak_x)):
                plt.plot(affect_side_peak_x[i], affect_side_peak_y[i], '*', markersize=10)  # 標出極值
                plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i])  # 標上極值數值
                # texta.append(plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i]))
                # print(texta)
            # adjust_text(texta)
            # plt.savefig('affect_side.png')
            plt.savefig(affect_pic_name)

            plt.clf()
            # plt.show()

            texts = []
            # x_peak = signal.find_peaks(yyy_L, distance=25)
            x_peak = AMPD(yyy_L)
            x_peak_list = []
            y_peak = []
            # print(x_peak[0])
            # print('the number of peaks is ' + str(len(x_peak[0])))
            plt.plot(xxx, yyy_L, label='Normal', color='black')
            plt.xlabel('time')
            plt.ylabel('knee angle')
            plt.legend()
            plt.title('Normal side (L)')
            # for i in range(len(x_peak[0])):
            #     x_peak_list.append(x_peak[0][i])
            #     y_peak.append(round(yyy_L[x_peak[0][i]], 2))

            for i in range(len(x_peak)):
                plt.plot(x_peak[i], yyy_L[x_peak[i]], '*', markersize=10)  # 標出極值
                # plt.text(x_peak[0][i], yyy_L[x_peak[0][i]], round(yyy_L[x_peak[0][i]], 2))  # 標上極值數值
                texts.append(plt.text(x_peak[i], yyy_L[x_peak[i]], round(yyy_L[x_peak[i]], 2)))
                y_peak.append(round(yyy_L[x_peak[i]], 2))
                x_peak_list.append(x_peak[i])

            # for i in range(len(x_peak[0])):
            #     plt.plot(x_peak[0][i], yyy_L[x_peak[0][i]], '*', markersize=10)  # 標出極值
            #     # plt.text(x_peak[0][i], yyy_L[x_peak[0][i]], round(yyy_L[x_peak[0][i]], 2))  # 標上極值數值
            #     texts.append(plt.text(x_peak[0][i], yyy_L[x_peak[0][i]], round(yyy_L[x_peak[0][i]], 2)))
            #     y_peak.append(round(yyy_L[x_peak[0][i]], 2))
            #     x_peak_list.append(x_peak[0][i])

            # plt.savefig('normal_side.png')
            plt.savefig(normal_pic_name)

            # plt.show()

            plt.plot(xxx, yyy_R, label='Affected', color='red')
            plt.xlabel('time')
            plt.ylabel('knee angle')
            plt.legend()
            plt.title('Combined')
            for i in range(len(affect_side_peak_x)):
                plt.plot(affect_side_peak_x[i], affect_side_peak_y[i], '*', markersize=10)  # 標出極值
                # plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i])  # 標上極值數值
                texts.append(plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i]))
            adjust_text(texts)
            plt.savefig(combine_pic_name)

            # plt.savefig('Combined.png')

        elif affect_side == '2':

            affect_side_peak_y = []
            affect_side_peak_x = []
            for i in range(len(step_list)):
                step_range = yyy_L[step_list[i][0]:step_list[i][1]]

                affect_side_peak_y.append(round(max(step_range), 2))  # 找出peak y值並存在lsit中 採計小數點後2位
                # affect_side_peak_x.append(np.where(step_range==max(step_range)))
                affect_side_peak_x.append(step_range.tolist().index(max(step_range)) + step_list[i][0])  # 找出peak x值並存在lsit中

                # pprint(step_range)
                # pprint(affect_side_peak_y)
                # pprint(affect_side_peak_x)

            plt.plot(xxx, yyy_L, label='Affected')
            plt.xlabel('time')
            plt.ylabel('knee angle')
            plt.legend()
            plt.title('Affected side (L)')
            for i in range(len(affect_side_peak_x)):
                plt.plot(affect_side_peak_x[i], affect_side_peak_y[i], '*', markersize=10)  # 標出極值
                plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i])  # 標上極值數值
            # plt.savefig('affect_side.png')
            plt.savefig(affect_pic_name)

            plt.clf()
            # plt.show()

            texts = []
            # x_peak1 = signal.find_peaks(yyy_R, prominence=1)
            x_peak = AMPD(yyy_R)
            print(x_peak)
            x_peak_list = []
            y_peak = []
            # print(x_peak[0])
            # print('the number of peaks is ' + str(len(x_peak[0])))
            plt.plot(xxx, yyy_R, label='Normal', color='black')
            plt.xlabel('time')
            plt.ylabel('knee angle')
            plt.legend()
            plt.title('Normal side (R)')
            for i in range(len(x_peak)):
                plt.plot(x_peak[i], yyy_R[x_peak[i]], '*', markersize=10)  # 標出極值
                # plt.text(x_peak[0][i], yyy_R[x_peak[0][i]], round(yyy_R[x_peak[0][i]], 2))  # 標上極值數值
                texts.append(plt.text(x_peak[i], yyy_R[x_peak[i]], round(yyy_R[x_peak[i]], 2)))
                y_peak.append(round(yyy_R[x_peak[i]], 2))
                x_peak_list.append(x_peak[i])


            # plt.savefig('normal_side.png')
            plt.savefig(normal_pic_name)

            # plt.show()

            plt.plot(xxx, yyy_L, label='Affected', color='red')
            plt.xlabel('time')
            plt.ylabel('knee angle')
            plt.legend()
            plt.title('Combined')
            for i in range(len(affect_side_peak_x)):
                plt.plot(affect_side_peak_x[i], affect_side_peak_y[i], '*', markersize=10)  # 標出極值
                # plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i])  # 標上極值數值
                texts.append(plt.text(affect_side_peak_x[i], affect_side_peak_y[i], affect_side_peak_y[i]))
            adjust_text(texts)
            # plt.savefig('Combined.png')
            plt.savefig(combine_pic_name)

        font_ = Font(
            name="標楷體",
            size=14,
            # italic=True,
            # color='ffff00',
            bold=False,
            strike=None
        )

        with pd.ExcelWriter(path=path_excel + "result" + str(walk) + ".xlsx", engine='openpyxl') as writer:
            sheet_name = output_sheet_name + str(walk)
            data.to_excel(writer, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            for row in worksheet.iter_rows(min_row=0, min_col=0):
                for cell in row:
                    cell.value = ""
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center', horizontal='center')
                    cell.font = font_
                    cell.border = openpyxl.styles.Border(left=None, right=None, top=None, bottom=None, diagonal=None,
                                                         diagonal_direction=0, outline=None, vertical=None, horizontal=None)

            worksheet['A1'] = "Normal side"
            worksheet['B1'] = "Affected side"
            worksheet.column_dimensions['A'].width = worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['A'].height = worksheet.column_dimensions['B'].height = 20
            for i in range(len(y_peak)):
                worksheet.cell(i + 2, 1).value = y_peak[i]
            for i in range(len(affect_side_peak_y)):
                worksheet.cell(i + 2, 2).value = affect_side_peak_y[i]
            img_normal = Image(normal_pic_name)
            img_affect = Image(affect_pic_name)
            img_combined = Image(combine_pic_name)
            worksheet.add_image(img_normal, 'E1')
            worksheet.add_image(img_affect, 'E26')
            worksheet.add_image(img_combined, 'Q1')
            plt.clf()
            walk += 1
# choose = input('退出程式輸入 2，繼續按Enter(更換完xlsx檔再按Enter):')
# if choose == '2':
#     ask_continue = 2
    output_file_name = os.getcwd() + '/' + ID + '/' + ID + '.xlsx'

    workbook = Workbook(path_excel + "result1.xlsx")
    for i in range(2, walk, 1):
        combine_xlsx_name = path_excel + "result" + str(i) + ".xlsx"
        workbook.combine(Workbook(combine_xlsx_name))
        # workbook.save(output_file_name)

    workbook.save(output_file_name)
    # workbook_final = Workbook(output_file_name)
    # workbook.worksheets.remove(workbook_final.worksheets['Evaluation Warning'])
    # workbook_final.worksheets.remove_at('Evaluation Warning')
    # workbook_final.save(output_file_name)
    workbook_final = openpyxl.load_workbook(output_file_name)
    del workbook_final['Evaluation Warning']
    # workbook_final.remove(workbook_final['Evaluation Warning'])
    workbook_final.save(output_file_name)
    choose = input('退出程式輸入 2，繼續按Enter(更換完xlsx檔再按Enter):')
    if choose == '2':
        ask_continue = 2

# choose = input('繼續該程式輸入1，退出輸入2:')
